import openpyxl
import pandas as pd
from openpyxl.styles import Font, colors, Alignment, PatternFill, Border, Side
import sys
from collections.abc import Iterable
from Clean_df import EDM
from Analytics import Analytics
from Read_json import Read_json


class Report(object):

    def __init__(self, df, campaign_id, template_path):
        '''
        :param df: EDM清洗好的dataframe对象，接口
        :param campaign_id: 这是一个执行类，所以要指明执行哪个campaign
        :param path:
        '''
        self.dataframe2 = df
        self.report_template_wb = openpyxl.load_workbook(template_path)
        self.report_template_ws = self.report_template_wb.active
        self.report = self.report_template_wb.active
        self.campaign_id = campaign_id
            
        self.main_end_row = None

        j = Read_json(campaign_id)
        data_dic = j.get_data_dic()
        self.main_click_list = j.get_main_click_list()
        self.main_click_list.insert(0,('Count', 'Main Link Name'))  #这样的好处是，在json里数据是干净好看的，做其他的事情时候拿到operation这边来加工
        self.other_click_list = j.get_other_click_list()
        self.other_click_list.insert(0,('Count', 'Other Link Name'))


        a = Analytics(self.get_campaign_id())

        self.analytics_list = a.data_package()
        self.analytics_list.insert(0,('Value', 'Attribute'))


    def get_dataframe2(self):
        return self.dataframe2

    def get_report_template_wb(self):
        return self.report_template_wb

    def get_report_template_ws(self):
        return self.report_template_ws

    def get_report(self):
        return self.report

    def get_campaign_id(self):
        return self.campaign_id

    def get_main_end_row(self):
        return self.main_end_row

    def get_main_click_list(self):
        return self.main_click_list

    def get_other_click_list(self):
        return self.other_click_list

    def get_analytics_list(self):
        return self.analytics_list

    def get_num(self, df, feature):
        id = self.get_campaign_id()
        return df[df['Campaign ID'] == id].loc[:, feature].iloc[-1]


    def report_dic(self):

        dic={}
        df = self.get_dataframe2()
        campaign_id = self.get_campaign_id()
        df_id = df[df['Campaign ID'] == campaign_id]

        feature_list = ['Sent', 'Delivered', 'Opened', 'Click', 'Unique Click', 'Launch Date']
        for feature in feature_list:
            if pd.isnull(df_id.loc[:, feature].iloc[-1]):
                dic[feature] = 0
            try:
                dic[feature] = int(df_id.loc[:,feature].iloc[-1])
            except:
                dic[feature] = df_id.loc[:,feature].iloc[-1]

        dic['Name'] = self.get_num(df, 'Campaign Name')
        dic['Campaign_id'] = campaign_id

        return dic


    def write_data(self, spider_dic={}, type='spider'):

        ws = self.get_report_template_ws()
        if type == 'spider':
            dic = spider_dic
        else:
            dic = self.report_dic()

         #这部分原样照搬上个方法
        df = self.get_dataframe2()
        campaign_id = self.get_campaign_id()
        df_id = df[df['Campaign ID'] == campaign_id]

        feature_list = ['Launch Date']
        for feature in feature_list:
            if pd.isnull(df_id.loc[:, feature].iloc[0]):
                dic[feature] = 0
            try:
                dic[feature] = int(df_id.loc[:, feature].iloc[0])
            except:
                dic[feature] = df_id.loc[:, feature].iloc[0]

        dic['Name'] = self.get_num(df, 'Campaign Name')
        dic['Campaign_id'] = campaign_id
        #-------------------------------#

        ws['A4'] = dic['Sent']
        ws['B4'] = dic['Delivered']
        try:
            ws['C4'] = dic['Opened']
        except KeyError:
            ws['C4'] = 0
        
        try:
            ws['D4'] = dic['Click']
        except KeyError:
            ws['D4'] = 0
            
        try:
            ws['E4'] = dic['Unique Click']
        except KeyError:
            ws['E4'] = 0

        ws['A1'] = dic['Name']
        ws['G1'] = dic['Campaign_id']
        ws['H1'] = dic['Launch Date']
        self.report = ws
        return ws

  

    def write_details(self, click_list, start_row = 8):
        '''
        这个方法是爬虫直接从网站上爬下来click的数据，也就是click_list作为接口，然后写在底下那里
        :param dic:
        :return:
        '''
        ws = self.get_report()
        num = len(click_list)
        end = start_row + num #计算结尾的行数，本应-1，但因为range函数最后一个不算，也就不减了，+1是因为start_row是表头的行号

        thin_border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))

        thick_border = Border(left=Side(border_style='thick', color='000000'),
                             right=Side(border_style='thick', color='000000'),
                             top=Side(border_style='thick', color='000000'),
                             bottom=Side(border_style='thick', color='000000'))

        title_fill = PatternFill('solid', fgColor="F0AB00")


        title_font = Font(name='Calibri', bold=True)

        for j in ['A', 'B', 'C', 'D', 'E']:  # 设置粗边框
            ws[j + str(start_row)].border = thick_border
            ws[j + str(start_row)].font = title_font
            ws[j + str(start_row)].fill = title_fill
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)


        for i in range(start_row, end):
            index = i - start_row
            ws['A'+str(i)] = click_list[index][1]  #因为列表已经按照倒序排列好了，所以直接按顺序排进来
            ws['E'+str(i)] = click_list[index][0]

            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)   #合并单元格
            ws['A' + str(i)].alignment = Alignment(horizontal='left', vertical='bottom')  #设置对齐方式：左对齐
            ws['E' + str(i)].alignment = Alignment(horizontal='left', vertical='bottom')

            if click_list[0][1] == 'Attribute':
                if i != start_row:
                    ws['A'+str(i)].number_format = '0.00%'
                    ws['E'+str(i)].number_format = '0.00%'

            for j in ['A','B','C','D','E']:     #设置边框
                ws[j + str(i)].border = thin_border

        self.main_end_row = end
        self.report = ws
        return ws


    def excute_report(self, data_dic,save_path = "\\\cnshag101.sha.global.corp.sap\\Restricted\\Marketing\\Marketing_Automation\\report\\", attribute=True):
        '''
        :param save_path: 指的是report产出的地址
        :return:
        '''

        main_click_list = self.get_main_click_list()
        other_click_list = self.get_other_click_list()
        analytics_list = self.get_analytics_list()
            
        wb = self.get_report_template_wb()
        self.write_data(spider_dic=data_dic, type='spider')
        self.write_details(main_click_list, start_row=8)
        self.write_details(other_click_list, start_row=self.get_main_end_row())
        if attribute:
            self.write_details(analytics_list, start_row=self.get_main_end_row())
    
        ws = self.get_report()
        dic = self.report_dic()
        path = save_path + dic['Name'] + '.xlsx'
        wb.save(path)
        return


class Tracker(object):

    def __init__(self, tracker_df, tracker_path):
        self.dataframe = tracker_df
        self.tracker_path = tracker_path
        self.tracker_template_wb = openpyxl.load_workbook(tracker_path)
        self.tracker_template_wb.active
        self.tracker_template_ws = self.tracker_template_wb['Tracker']
        self.col_dic = {'Sent': 'R',
                        'Delivered': 'S',
                        'Opened': 'T',
                        'Soft Bounces': 'U',
                        'Hard Bounces': 'V',
                        'Click': 'W',
                        'Unique Click': 'X',
                        'Communication Limits Reached': 'Y',
                        }

    def get_dataframe(self):
        return self.dataframe

    def get_tracker_template_wb(self):
        return self.tracker_template_wb

    def get_tracker_template_ws(self):
        return self.tracker_template_ws

    def get_col_dic(self):
        return self.col_dic

    def get_tracker_path(self):
        return self.tracker_path

    def catch_row_num(self, campaign_id):
        '''
        找到这个campaign_id在excel中的行数
        return:返回第一个含有此campaign_id的excel行数
        '''
        
        df = self.get_dataframe()

        index = df[df['Campaign ID'] == campaign_id].index[0]
        row_num = index + 2

        return row_num

    def write_data(self, data_dic, campaign_id):
        '''
        param data_dic 爬虫抓出来的面板数据
        return  ws，填好数据后的tracker
        '''
        
        ws = self.get_tracker_template_ws()
        col_dic = self.get_col_dic()
        row_num = self.catch_row_num(campaign_id)
        for item in col_dic:
            try:
                ws[col_dic[item] + str(row_num)] = data_dic[item]
            except KeyError:   #这个地方是为了避免出现两个词典，key不一样的情况，比如我有communication limit这个key，但爬下来的数据是没有的
                continue

        self.tracker_template_ws = ws
        return

    def write_campaign_id(self, index, campaign_id):
        '''
        给定行号，写入campaign id
        '''
        ws = self.get_tracker_template_ws()
        row_num = index + 2
        ws['K'+str(row_num)] = int(campaign_id)

        return


    def save_data(self):
        path = self.get_tracker_path()
        wb = self.get_tracker_template_wb()
        wb.save(path)
        return


if __name__ == '__main__':
    campaign_id = 5509
    # df = pd.read_excel(r'C:\Users\C5293427\Desktop\MA\Request_Tracker.xlsx', sheet_name='Tracker', encoding = 'utf-8')
    tracker_path = r'C:\Users\C5293427\Desktop\MA\Request_Tracker.xlsx'
    report_template = r'C:\Users\C5293427\Desktop\MA\report\Report_template.xlsx'
    report_save = 'C:/Users/C5293427/Desktop/MA/report/'
    edm = EDM(tracker_path)
    df1 = edm.get_dataframe1()

    # spider = DataSpider(campaign_id)
    # spider.scratch_data()

    # main_click_list = spider.get_main_click_list()
    # other_click_list = spider.get_other_click_list()
    # data_dic = spider.get_data_dic()

    r = Report(df=df1, campaign_id=campaign_id, template_path=report_template)
    print(r.get_main_click_list())
    

    # t = Tracker(tracker_df=df, tracker_path=tracker_path)
    # t.write_campaign_id(561,campaign_id)
    # t.save_data()












