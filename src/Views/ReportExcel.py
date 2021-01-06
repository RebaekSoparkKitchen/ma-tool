"""
@Description: 此类负责report的excel表格的内容及排版
@Author: FlyingRedPig
@Date: 2020-05-08 11:35:14
@LastEditors: FlyingRedPig
@LastEditTime: 2020-09-03 10:37:13
@FilePath: \MA_tool_v2\src\Views\ReportExcel.py
"""
import sys

sys.path.append("../..")
from src.Connector.MA import MA
from src.Models.Report.ReportData import ReportData
from src.Models.Report.ClickPerformance import ClickPerformance
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows


class ReportExcel(object):
    def __init__(self, smc_campaign_id):
        template_path = MA().read_config()['file_location']['report_template']
        self.smc_campaign_id = smc_campaign_id
        self.report_wb = openpyxl.load_workbook(template_path)
        self.report_ws = self.report_wb.active
        self.report_data = ReportData(smc_campaign_id)
        self.table_width = 5
        self.save_path = MA().read_config()['file_location']['report_save']

    @staticmethod
    def standard_font() -> Font:
        return Font(name='Calibri', bold=False, size=11)

    @staticmethod
    def header_font() -> Font:
        return Font(name='Calibri', bold=True, size=11)

    @staticmethod
    def number_font() -> Font:
        return Font(name='Calibri', bold=True, size=14, color="00aecc")

    def set_format(self) -> None:
        """
        @description: 在读取template之后，先整体对excel表格进行一些font,border的设置
        @param {type} 
        @return: 
        """
        border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000'))

        for col in self.report_ws.columns:
            if col == list(self.report_ws.columns)[-1]:
                break
            for i in col:
                i.border = border
                i.font = self.standard_font()
        return

    def add_name(self) -> None:
        self.report_ws['A1'] = self.report_data.campaign_name
        self.report_ws['A1'].font = self.header_font()
        return

    def __add_overview(self) -> None:
        """
        @description: 添加overview块的字符串。ps.这边的接口设计是，所有计算均在对应的RequestTracker和LocalData类中完成，以降低耦合度。
        @param {type} 
        @return: 
        """
        report_data = self.report_data
        basic_data = self.report_data.basic_performance
        execution_time = report_data.blast_date

        # calculation, transfer float to percentage
        open_rate = "%.2f%%" % (basic_data.open_rate * 100)
        ctr = "%.2f%%" % (basic_data.ctr * 100)
        click_to_open = "%.2f%%" % (basic_data.unique_click_to_open_rate * 100)

        # insert data to the template sentence
        time_str = f"\n- Date of execution: 08:00 AM on {execution_time}"
        open_rate_str = f"\n- Open Rate - {open_rate} （GC benchmark: 4-6%）"
        ctr_str = f"\n- CTR - {ctr}  (GC benchmark: 0.2 – 0.4 %)"
        click_to_open_str = f"\n- Click-to-open Rate - {click_to_open} (GC benchmark: 5 - 7 % )"

        self.report_ws['A2'] = 'Summary:' + time_str + open_rate_str + ctr_str + click_to_open_str
        return

    def __add_basic_performance(self) -> None:
        """
        @description: 添加基本表现信息，这些信息都是从localdata得到的，所以特意从localdata做了接口传过来。
        @param {type} 
        @return: 
        """
        basic_data = self.report_data.basic_performance

        self.report_ws['A4'] = int(basic_data.sent)
        self.report_ws['B4'] = int(basic_data.delivered)
        self.report_ws['C4'] = int(basic_data.opened)
        self.report_ws['D4'] = int(basic_data.click)
        self.report_ws['E4'] = int(basic_data.unique_click)

        # 设置 header 格式
        basic_performance_header = [3, 6]
        for row in basic_performance_header:
            for col in range(1, self.table_width + 1):
                self.report_ws.cell(row=row, column=col).font = self.header_font()

        # 设置 data 格式
        basic_performance_number = [4, 7]
        for row in basic_performance_number:
            for col in range(1, self.table_width + 1):
                self.report_ws.cell(row=row, column=col).font = self.number_font()

        return

    def __add_click_performance(self) -> None:
        """
        @description: 将点击情况写入表格，来源于LocalData的两个方法，直接提取df过来，这里主要是一些格式上的调整。 
        @param {type} 
        @return: 
        """
        ws = self.report_ws

        main_links = ClickPerformance.main_link_list(self.smc_campaign_id)
        other_links = ClickPerformance.other_link_list(self.smc_campaign_id)
        start_row = 9
        table_width = 5
        merge_col_numbers = 3
        current_row = start_row

        # 将数据导入
        ws['A' + str(current_row)] = 'Main Link Name'
        ws['E' + str(current_row)] = 'Click Numbers'

        for link in main_links:
            current_row += 1
            ws['A' + str(current_row)] = link.link_name
            ws['E' + str(current_row)] = link.click_number

        current_row += 1

        ws['A' + str(current_row)] = 'Other Link Name'
        ws['E' + str(current_row)] = 'Click Numbers'

        for link in other_links:
            current_row += 1
            ws['A' + str(current_row)] = link.link_name
            ws['E' + str(current_row)] = link.click_number

        # link_name = iter(['Main Link Name', 'Other Link Name'])
        # for df in [mainDf, otherDf]:
        #     name = next(link_name)
        #     for i in range(merge_col_numbers - 1):
        #         df[None] = None
        #     if df.empty:
        #         df[name] = None
        #         df['Click Numbers'] = None
        #
        #     df = df[[name, None, None, None, 'Click Numbers']]
        #     for r in dataframe_to_rows(df, index=False, header=True):
        #         self.report_ws.append(r)


        # 合并单元格
        click_length = len(self.report_data.click_performance)
        for i in range(start_row, start_row + click_length + 2):  # 3 指两个df的header+range函数不计算最后一位
            self.report_ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=table_width - 1)

        # 加颜色/字体加粗
        fill = PatternFill("solid", fgColor='f0b8b8')

        for col in range(1, table_width + 1):
            self.report_ws.cell(row=start_row, column=col).fill = fill
            self.report_ws.cell(row=start_row, column=col).font = self.header_font()
            self.report_ws.cell(row=start_row + len(main_links) + 1, column=col).fill = fill  # 多一个加1是加在header上面
            self.report_ws.cell(row=start_row + len(main_links) + 1, column=col).font = self.header_font()  #
            # 多一个加1是加在header上面

        # 设置边框和对齐方式
        border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000'))
        align = Alignment(horizontal='left', vertical='center', wrap_text=True)

        for row in range(start_row, start_row + len(main_links) + len(other_links) + 2):
            for col in range(1, table_width + 1):
                self.report_ws.cell(row=row, column=col).border = border
                self.report_ws.cell(row=row, column=col).alignment = align

        return

    def __create(self, path: str) -> None:
        """
        @description: 关于保存过程的封装，这里值得注意要在permission error时提示关闭。 
        @param {str} excel 表格的存储地址，只需写到父级目录，excel的名字会自动从tracker中获取。 
        @return: 
        """
        while True:
            # 文件名中出现特殊符号，要替换掉
            name = self.report_data.campaign_name
            name = name.replace(r'/', r'&')
            name = name.replace(' ', '_')
            name = name.replace('|', '')
            try:
                self.report_wb.save(path + '{}.xlsx'.format(name.strip()))
                break
            except PermissionError:
                reminder = input('请关闭{}的数据报告，是否现在重试？(y/n)'.format(name))
                if reminder.lower() == 'y':
                    continue
                elif reminder.lower() == 'n':
                    print('由于您执意不关闭{}的数据报告，我们无法保存，抱歉！'.format(name))
                    break
                else:
                    print('请输入正确的命令(y/n)')
        return

    def save(self, path: str = "") -> None:
        """
        @description: 对外唯一接口，创建一个report excel表格。 
        @param {str} path: excel表格的存储地址，只需写到父级目录，excel的名字会自动从tracker中获取。  
        @return: 
        """
        self.set_format()
        self.add_name()
        self.__add_overview()
        self.__add_basic_performance()
        self.__add_click_performance()
        if path:
            self.__create(path)
        else:
            self.__create(self.save_path)
        return


if __name__ == "__main__":
    import time
    t1= time.time()
    r = ReportExcel(6414)
    r.save('C:/Users/C5293427/Desktop/test/report_excel/')
    t2= time.time()
    print(t2 - t1)
