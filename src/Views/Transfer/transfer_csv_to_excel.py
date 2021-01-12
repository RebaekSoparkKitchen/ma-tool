import pandas as pd
import csv
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill
from src.Connector.MA import MA


class ExportedData(MA):

    def __init__(self, path):
        """
        初始化，self.df用来导入原始文件，self.clean_df用来存储清洗后的文件
        """
        super().__init__()
        self.df = pd.read_csv(path, sep='^', error_bad_lines=False,
                              quoting=csv.QUOTE_NONE, engine="python", encoding="utf_8_sig")

        self.path = path
        self.clean_df = 0  # Placeholder

    def get_data(self):
        return self.df

    def get_path(self):
        return self.path

    def get_clean_data(self):
        return self.clean_data

    def clean_data(self):
        """
        :param self: 导入原始文件
        :return: 清洗后的数据，将电话号码转为str，并将清洗后的数据存入self.clean_df
        """
        df = self.get_data()

        for i in ['Acc Phone Number', 'CP Phone Number', 'CP Mobile Phone Number', 'Opt-in Phone']:

            try:
                # 尝试先对他们整数化，如若不行，eg:+85228880888|+85263680340，则直接str()
                df[i] = df[i].apply(lambda x: str(
                    int(x)) if pd.notnull(x) else x)
            except ValueError:
                df[i] = df[i].apply(lambda x: str(
                    x) if pd.notnull(x) else x)

        self.clean_df = df

        return df

    def read_file(self, path):
        df = pd.read_csv(open((path), 'r', encoding='utf-8'), sep=';')
        return df

    def add_color(self, click_file_path, match_attribute='DB_key'):
        """
        为点击的人添加颜色
        color_file是一个df文件
        """
        if self.get_clean_data() == 0:
            self.clean_data()
        df = self.get_clean_data()
        click_file = self.read_file(click_file_path)

        row_num = []
        for i in range(len(df)):
            if df.loc[i, match_attribute] in list(click_file[match_attribute]):
                row_num.append(i+2)

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        ws.insert_cols(2)
        ws.cell(row=1, column=2, value='If_Click')

        orange_fill = PatternFill(fill_type='solid', fgColor="FFC125")

        for i in row_num:
            ws.cell(row=i, column=2, value='True')
            for j in range(1, len(df)+10):
                ws.cell(row=i, column=j).fill = orange_fill

        border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))

        ws['B1'].border = border
        ws['B1'].font = Font(bold=True)

        wb.save(self.get_path())
        return

    def to_xlsx(self, path):
        """
        储存文件
        :param self: 对象
        :param path: 文件存储路径
        :return:
        """
        self.clean_data()
        self.clean_df.to_excel(path, encoding='utf_8_sig', index=False)

        return


class Highlight(object):

    def __init__(self, path):
        self.path = path
        self.df = pd.read_excel(path, encoding='utf-8')

    def get_path(self):
        return self.path

    def get_df(self):
        return self.df

    def add_color(self, click_file_path, match_attribute='Email'):
        """
        为点击的人添加颜色
        color_file是一个df文件
        """
        path = self.get_path()
        df = self.get_df()
        click_file = pd.read_csv(click_file_path, sep=',')

        row_num = []
        for i in range(len(df)):
            if df.loc[i, match_attribute] in list(click_file[match_attribute]):
                row_num.append(i+2)

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        ws.insert_cols(2)
        ws.cell(row=1, column=2, value='If_Click')

        orange_fill = PatternFill(fill_type='solid', fgColor="FFC125")

        for i in row_num:
            ws.cell(row=i, column=2, value='True')
            for j in range(1, len(df)+10):
                ws.cell(row=i, column=j).fill = orange_fill

        border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))

        ws['B1'].border = border
        ws['B1'].font = Font(bold=True)

        wb.save(self.get_path())
        print('Highlight successfully!')
        return


if __name__ == "__main__":
    e = Highlight(
        r'C:/Users/C5293427/Desktop/Data_exported/成品/CN_智能ERP_open.xlsx')
    e.add_color(
        r'C:/Users/C5293427/Desktop/Data_exported/csv_file/CN_智能ERP_click_email.csv')
