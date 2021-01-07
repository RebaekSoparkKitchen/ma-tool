"""
@Description: SimpleTracker类也是RequestTracker类的子类，它所提供的服务非常简单：(1) 提供一个日历形式的dataframe; (2) 提供一个日历形式的excel表格
@Author: FlyingRedPig
@Date: 2020-04-30 18:03:27
@LastEditors: FlyingRedPig
@LastEditTime: 2020-08-19 10:59:07
@FilePath: \MA_tool\src\Tracker\SimpleTracker.py
"""
import datetime as dt
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


class CalendarExcel(object):

    def __init__(self, data):
        self.data = data
        self.title = "Campaign Calender"

    def set_title(self, ws):
        ws.title = self.title

    @staticmethod
    def set_data(ws, data):
        for r in data:
            ws.append(r)

    @staticmethod
    def set_header(ws):
        headers = ['Launch Date', 'Weekday', 'Campaign Name', 'Event Date', 'Owner', 'Campaign ID']
        ws.append(headers)

    @staticmethod
    def set_col_width(ws):
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 15

    @staticmethod
    def set_font(ws):
        eng_font = Font(name='Cambria', size=11, bold=False)
        cn_font = Font(name=u'微软雅黑', size=11, bold=False)
        first_row_font = Font(name='Times New Roman', size=11, bold=True)

        for col in ws.columns:
            for i in col:
                i.font = eng_font

        for i in ws['1']:
            i.font = first_row_font

    @staticmethod
    def set_align(ws):

        align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        for col in ws.columns:
            for i in col:
                i.alignment = align

    @staticmethod
    def set_border(ws):
        border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000'))

        for col in ws.columns:
            for i in col:
                i.border = border

    @staticmethod
    def date_range(end: dt.date, diff: int) -> list:
        start = end - dt.timedelta(days=diff)
        dates = []
        for i in range(diff):
            start += dt.timedelta(days=1)
            dates.append(start.strftime('%Y-%m-%d'))
        return dates

    @staticmethod
    def set_color(ws):

        color_list = ['aecdc2', 'f0b8b8']
        color = 'aecdc2'

        def switch(color, color_list):

            for i in color_list:
                if i == color:
                    continue
                return i

        for i in ws.rows:
            for j in i:
                j.fill = PatternFill('solid', fgColor=color)
            if i[1].value == '星期日':
                color = switch(color, color_list)

        pattern2 = PatternFill('solid', fgColor='f3e8c2')  # header颜色
        for i in ws['1']:
            i.fill = pattern2

        # 开始做今天起倒退五天的渐进色

        def one_day_color(ws, color, date):

            pattern3 = PatternFill('solid', fgColor=color)
            for i in ws.rows:
                if isinstance(i[0].value, str) or (i[0].value is None):
                    continue
                else:
                    if i[0].value.strftime('%Y-%m-%d') == str(date):
                        for j in i:
                            j.fill = pattern3

        today = dt.date.today()
        traffic_range = CalendarExcel.date_range(today, 5)  # 整体转换为列表类型
        # traffic_range = list(map(lambda x: x.date(), traffic_range))  # 元素转换为date类型
        gradient_color = ['ff6e54', 'ffa600', 'ffa600', 'ffa600', 'ffa600']
        color_dict = dict(zip(traffic_range, gradient_color))  # 将日期和颜色对应成字典
        for date in color_dict.keys():
            one_day_color(ws, color_dict[date], date)

    @staticmethod
    def set_merge(ws):

        col = ws['A']
        date_dic = {}
        index = 0
        for i in col:
            index += 1
            item = i.value
            if item not in date_dic.keys():
                date_dic[item] = [index]
            else:
                date_dic[item].append(index)

        def merge(index_list):
            ws.merge_cells(
                start_row=min(index_list),
                start_column=1,
                end_row=max(index_list),
                end_column=1)

            ws.merge_cells(
                start_row=min(index_list),
                start_column=2,
                end_row=max(index_list),
                end_column=2)
            return

        # 去除所有不需要合并的对象
        index = list(date_dic.values())
        trash = []
        for i in index:
            if len(i) == 1:
                trash.append(i)

        for i in trash:
            index.remove(i)

        list(map(merge, index))  # 对每个日期都合并单元格

        return

    @staticmethod
    def set_freeze(ws):
        ws.freeze_panes = 'A2'

    def save(self, path):

        wb = Workbook()
        ws = wb.active
        self.set_title(ws)
        ws = wb[self.title]
        self.set_header(ws)
        self.set_data(ws, self.data)
        self.set_col_width(ws)  # 0s
        self.set_font(ws)  # 0.018s
        self.set_align(ws)  # 0.007s
        self.set_border(ws)  # 0.037s
        self.set_color(ws)  # 0.07s
        self.set_merge(ws)  # 0.14s
        self.set_freeze(ws)
        while True:
            try:
                wb.save(path)  # 0.03s
                print('simple tracker已创建成功')
                break
            except PermissionError:
                answer = input("simple_tracker.xlsx已打开，请关闭后重试，谢谢(y/n)")
                if answer.lower() == 'y':
                    pass
                elif answer.lower() == 'n':
                    print('由于您执意不关闭simple_tracker.xlsx，我们只好不干活了，simple tracker 没有创建~')
                    break
                else:
                    print('请输入正确的命令(y/n)')
        return
