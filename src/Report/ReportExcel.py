"""
@Description: 此类负责report的excel表格的内容及排版
@Author: FlyingRedPig
@Date: 2020-05-08 11:35:14
@LastEditors: FlyingRedPig
@LastEditTime: 2020-09-03 10:37:13
@FilePath: \MA_tool\src\Report\ReportExcel.py
"""
import sys
sys.path.append("../..")
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import openpyxl
from src.LocalDataBase.LocalData import LocalData
from src.Tracker.Analytics import Analytics
from src.Connector.MA import MA
from openpyxl.utils.dataframe import dataframe_to_rows


class ReportExcel(MA):

    def __init__(
        self,
        campaignId):
        super().__init__()
        templatePath = self.readConfig()['file_location']['reportTemplate']
        self.reportWb = openpyxl.load_workbook(templatePath)
        self.reportWs = self.reportWb.active
        self.campaignId = campaignId
        self.trackerData = Analytics(self.getCampaignId())
        self.localData = LocalData()
        self.tableWidth = 5
        self.savePath = self.readConfig()['file_location']['reportSave']
            

    def getCampaignId(self) -> int:
        return int(self.campaignId)

    def getTrackerData(self) -> Analytics:
        return self.trackerData

    def getLocalData(self) -> LocalData:
        return self.localData

    def getTableWidth(self) -> int:
        return self.tableWidth

    def getSavePath(self) -> str:
        return self.savePath

    def standardFont(self) -> Font:
        return Font(name='Calibri', bold=False, size=11)

    def headerFont(self) -> Font:
        return Font(name='Calibri', bold=True, size=11)
    
    def numberFont(self) -> Font:
        return Font(name='Calibri', bold=True, size=14, color="00aecc")

    def setFormat(self) -> None:
        """
        @description: 在读取template之后，先整体对excel表格进行一些font,border的设置
        @param {type} 
        @return: 
        """
        border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000'))
        
        for col in self.reportWs.columns:
            if col == list(self.reportWs.columns)[-1]:
                break
            for i in col:
                i.border = border
                i.font = self.standardFont()
        return


    def __addName(self) -> None:
        trackerData = self.getTrackerData()
        self.reportWs['A1'] = trackerData.name()
        self.reportWs['A1'].font = self.headerFont()
        return

    def __addOverview(self) -> None:
        """
        @description: 添加overview块的字符串。ps.这边的接口设计是，所有计算均在对应的RequestTracker和LocalData类中完成，以降低耦合度。
        @param {type} 
        @return: 
        """
        trackerData = self.getTrackerData()
        localData = self.getLocalData()

        executionTime = trackerData.executionTime()

        openRate = "%.2f%%" % (
            localData.metricOpenRate(self.getCampaignId()) * 100)
        ctr = "%.2f%%" % (localData.metricCTR(self.getCampaignId()) * 100)
        clickToOpen = "%.2f%%" % (
            localData.metricClickToOpen(self.getCampaignId()) * 100)

        timeStr = "\n- Date of execution: 08:00 AM on {}".format(executionTime)
        openRateStr = "\n- Open Rate - {} （GC benchmark: 4-6%）".format(openRate)
        ctrStr = "\n- CTR - {}  (GC benchmark: 0.2 – 0.4 %)".format(ctr)
        clickToOpenStr = "\n- Click-to-open Rate - {} (GC benchmark: 5 - 7 % )".format(clickToOpen)
        self.reportWs['A2'] = 'Summary:' + timeStr + openRateStr + ctrStr + clickToOpenStr

        return

    def __addBasicPerformance(self) -> None:
        """
        @description: 添加基本表现信息，这些信息都是从localdata得到的，所以特意从localdata做了接口传过来。
        @param {type} 
        @return: 
        """
        LocalData = self.getLocalData()

        self.reportWs['A4'] = LocalData.metricSent(self.getCampaignId())
        self.reportWs['B4'] = LocalData.metricDeliver(self.getCampaignId())
        self.reportWs['C4'] = LocalData.metricOpen(self.getCampaignId())
        self.reportWs['D4'] = LocalData.metricClick(self.getCampaignId())
        self.reportWs['E4'] = LocalData.metricUniqueClick(self.getCampaignId())

        #设置 header 格式
        basicPerformanceHeader = [3,6]
        for row in basicPerformanceHeader:
            for col in range(1,self.getTableWidth()+1):
                self.reportWs.cell(row=row, column=col).font = self.headerFont()
        
        #设置 data 格式
        basicPerformanceNumber = [4,7]
        for row in basicPerformanceNumber:
            for col in range(1,self.getTableWidth()+1):
                self.reportWs.cell(row=row, column=col).font = self.numberFont()

        return

    def __addClickPerformance(self) -> None:
        """
        @description: 将点击情况写入表格，来源于LocalData的两个方法，直接提取df过来，这里主要是一些格式上的调整。 
        @param {type} 
        @return: 
        """
        mainDf = self.getLocalData().mainClickPerformanceDf(self.getCampaignId())
        otherDf = self.getLocalData().otherClickPerformanceDf(self.getCampaignId())

        #将数据导入
        mergeColNumbers = 3
        linkName = iter(['Main Link Name','Other Link Name'])
        for df in [mainDf, otherDf]:
            name = next(linkName)
            for i in range(mergeColNumbers-1):
                df[None] = None
            if df.empty:
                df[name] = None
                df['Click Numbers'] = None
            
            df = df[[name,None,None,None,'Click Numbers']]
            for r in dataframe_to_rows(df, index=False, header=True):
                self.reportWs.append(r)

        clickStartRow = 8 
        tableWidth = 5

        #合并单元格
        self.reportWs.merge_cells(start_row=clickStartRow, start_column=1, end_row=clickStartRow, end_column=5)
        clickLength = len(mainDf) + len(otherDf)
        for i in range(clickStartRow,clickStartRow + clickLength + 3): # 3 指两个df的header+range函数不计算最后一位 
            self.reportWs.merge_cells(start_row=i, start_column=1, end_row=i, end_column= tableWidth-1)

        #加颜色/字体加粗
        fill = PatternFill("solid", fgColor='f0b8b8')
       
        for col in range(1,tableWidth+1):
            self.reportWs.cell(row=clickStartRow+1,column=col).fill = fill
            self.reportWs.cell(row=clickStartRow+1,column=col).font = self.headerFont()
            self.reportWs.cell(row=clickStartRow+len(mainDf)+1+1,column=col).fill = fill #多一个加1是加在header上面
            self.reportWs.cell(row=clickStartRow+len(mainDf)+1+1,column=col).font = self.headerFont() #多一个加1是加在header上面
        
        #设置边框和对齐方式
        border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000'))
        align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        for row in range(clickStartRow,clickStartRow + len(mainDf) + len(otherDf) +3):
            for col in range(1,tableWidth+1):
                self.reportWs.cell(row=row, column=col).border = border
                self.reportWs.cell(row=row, column=col).alignment = align

        return

    

    def __save(self, path:str) -> None :
        """
        @description: 关于保存过程的封装，这里值得注意要在permission error时提示关闭。 
        @param {str} excel 表格的存储地址，只需写到父级目录，excel的名字会自动从tracker中获取。 
        @return: 
        """
        while True:
            try:
                # 文件名中出现特殊符号，要替换掉
                name = self.getTrackerData().name().replace(r'/', r'&')
                name = self.getTrackerData().name().replace(' ', '_')
                name = self.getTrackerData().name().replace('|', '')
                self.reportWb.save(path+'{}.xlsx'.format(name.strip()))
                break
            except PermissionError:
                name = self.getTrackerData().name()
                reminder = input('请关闭{}的数据报告，是否现在重试？(y/n)'.format(name))
                if reminder.lower() == 'y':
                    continue
                elif reminder.lower() == 'n':
                    print('由于您执意不关闭{}的数据报告，我们无法保存，抱歉！'.format(name))
                    break
                else:
                    print('请输入正确的命令(y/n)')
        return

    def create(self, path:str="") -> None:
        """
        @description: 对外唯一接口，创建一个report excel表格。 
        @param {str} path: excel表格的存储地址，只需写到父级目录，excel的名字会自动从tracker中获取。  
        @return: 
        """
        self.setFormat()
        self.__addName()
        self.__addOverview()
        self.__addBasicPerformance()
        self.__addClickPerformance()
        self.__save(self.getSavePath())
        return 


if __name__ == "__main__":
    r = ReportExcel(10149)
    r.addClickPerformance()

