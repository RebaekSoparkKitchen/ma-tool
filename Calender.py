import openpyxl
import pandas as pd
import datetime as dt
from openpyxl.styles import Font, colors, Alignment, Border, Side, PatternFill
from time import time
from Clean_df import EDM
import numpy as np



class Calender(object):
    '''
    生成一个按时序排列的表格
    '''
    def __init__(self, df, save_path):
        '''

        :param df: 从EDM clean_date抓出来一个df，这是来源，输入
        :param save_path: 这是输出，搞一个地址，把输出存进去
        '''

        self.dataframe2 = df
        self.path = save_path
        self.start_date = dt.date.today() + dt.timedelta(days=-30)

    def get_path(self):
        return self.path

    def get_dataframe2(self):
        return self.dataframe2

    def time_based_data(self):

        df = self.get_dataframe2()
        df1 = df[['Launch Date', 'weekday', 'Campaign Name', 'Event Date', 'Owner ']]
        df1.sort_values(by="Launch Date", ascending=True, inplace=True)
        return df1

    def add_start_date(self, start_date = '20200101'):
        '''

        :param start_date: 表格从哪个日期开始
        :return: 重新排序后的数据
        '''
        start_date = self.start_date
        if type(start_date) == int:
            date_str = str(start_date)
        df = self.time_based_data()
        #date = parse(start_date).date()
        df = df[df['Launch Date'] > start_date]  #按时间筛选出start_date以后的时间
        df = df.reset_index(drop = True)  #重新索引

        return df

    def get_date_range(self):
        #找到从开始到现在所有的日期，若不存在，则插值，重新排序
        df = self.add_start_date()
        df['Launch Date'].dropna(inplace=True)
        df['Event Date'].dropna(inplace=True)
        launch_date_max = df['Launch Date'][df['Launch Date'].apply(lambda x: isinstance(x, dt.date))].max()
        launch_date_min = df['Launch Date'][df['Launch Date'].apply(lambda x: isinstance(x, dt.date))].min()
        event_date_max = df['Event Date'].max()
        event_date_min = df['Event Date'].min()
        start_date = df['Launch Date'].min()

        if launch_date_max > event_date_max:
            end_date = event_date_max
        else:
            end_date = launch_date_max

        start_date_str = str(start_date)
        end_date_str = str(end_date)

        date_index = pd.date_range(start_date_str, end_date_str)

        return date_index

    #从这里开始补充其他日期

    def create_one_line_dataframe(self):

        df = pd.DataFrame(
            {'Launch Date': [np.nan], 'weekday': [np.nan], 'Campaign Name': [np.nan], 'Event Date': [np.nan],
             'Owner ': [np.nan]})

        return df

    def add_date_to_one_line_dataframe(self,date):

        df = self.create_one_line_dataframe()
        df['Launch Date'] = date

        return df

    def add_other_date_in_dataframe(self):

        date_range = self.get_date_range()
        df = self.add_start_date()

        for i in date_range:

            if i.date() not in list(df['Launch Date']):

                df_new = self.add_date_to_one_line_dataframe(i.date())
                df = df.append(df_new)

        dic = {
            0: '星期一',
            1: '星期二',
            2: '星期三',
            3: '星期四',
            4: '星期五',
            5: '星期六',
            6: '星期日',
        }
        df['weekday'] = df['Launch Date']
        df['weekday'] = df['weekday'].apply(lambda x: dic[x.weekday()])

        df.sort_values(by="Launch Date", ascending=True, inplace=True)
        df = df.reset_index(drop=True)  # 重新索引

        return df

    def to_xlsx(self):

        df = self.add_other_date_in_dataframe()
        df1 = df.rename(columns={"weekday": "Weekday"})
        df1.to_excel(self.get_path(), index = False)
        return


class Simple_tracker(object):

    def __init__(self):
        self.wb = openpyxl.load_workbook(r'C:\Users\C5293427\Desktop\MA\Simple_Tracker_V3.xlsx')
        self.ws = self.wb.active

    def get_ws(self):
        return self.ws

    def get_wb(self):
        return self.wb

    def set_column_width(self):
        '''
        调节列宽
        :return:
        '''
        ws = self.get_ws()
        dic = {0: 'A', 1: 'B', 2: 'C', 3: 'D', 4: 'E', 5: 'F'}

        # 获取每一列的内容的最大宽度
        i = 0
        col_width = []
        # 每列
        for col in ws.columns:
            # 每行
            for j in range(len(col)):
                if j == 0:

                    # 数组增加一个元素
                    col_width.append(len(str(col[j].value)))
                else:
                    # 获得每列中的内容的最大宽度
                    if col_width[i] < len(str(col[j].value)):
                        col_width[i] = len(str(col[j].value))
            i = i + 1

        # 设置列宽
        for i in range(len(col_width)):
            # 根据列的数字返回字母
            col_letter = dic[i]
            ws.column_dimensions[col_letter].width = col_width[i] + 8

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 25

        return

    def set_font(self):
        '''
        设置字体
        :return:
        '''
        ws = self.get_ws()
        Roman_font = Font(name='Times New Roman', size=11, bold=False)
        Yahei_font = Font(name=u'微软雅黑', size=11, bold=False)
        first_row_font = Font(name='Times New Roman', size=11, bold=True)

        for i in ws['C']:
            i.font = Yahei_font

        for col in ws.columns:
            for i in col:
                i.font = Roman_font

        for i in ws['B']:
            i.font = Yahei_font

        for i in ws['1']:
            i.font = first_row_font

        return

    def set_align(self):
        '''
        设置所有左对齐
        :return:
        '''
        ws = self.get_ws()
        align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        for col in ws.columns:
            for i in col:
                i.alignment = align
        return

    def set_border(self):
        '''
        设置边框
        :return:
        '''
        ws = self.get_ws()
        border = Border(left=Side(border_style='thin', color='000000'),

                        right=Side(border_style='thin', color='000000'),

                        top=Side(border_style='thin', color='000000'),

                        bottom=Side(border_style='thin', color='000000'))

        for col in ws.columns:
            for i in col:
                i.border = border
        return

    def switch(self,color, color_list):
        '''
        set_color的辅助函数
        :param color_list:
        :return:
        '''
        for i in color_list:
            if i == color:
                continue
            return i

    def set_color(self):
        '''
        设置背景颜色
        :return:
        '''
        ws = self.get_ws()
        color_list = ['90EE90', 'CCCCCC']
        color = '90EE90'

        for i in ws.rows:
            for j in i:
                j.fill = PatternFill('solid', fgColor=color)
            if i[1].value == '星期日':
                color = self.switch(color, color_list)

        pattern2 = PatternFill('solid', fgColor='f0ab00')

        for i in ws['1']:
            i.fill = pattern2
            
        pattern3 = PatternFill('solid', fgColor='ffaf00')
        
        today = dt.date.today()
        
        for i in ws.rows:
            if type(i[0].value) == str:
                continue
            else:
                if i[0].value.strftime('%Y-%m-%d') == str(today):
                    for j in i:
                        j.fill = pattern3

        return

    def create_excel(self):
        '''
        将所有的设置整合到一起
        :return:
        '''
        ws = self.get_ws()
        wb = self.get_wb()
        self.set_column_width()
        self.set_font()
        self.set_align()
        self.set_border()
        self.set_color()
        wb.save(r'C:\Users\C5293427\Desktop\MA\Simple_Tracker_V3.xlsx')

if __name__ == '__main__':
    t_path = r'C:\Users\C5293427\Desktop\MA\Request_Tracker20190523.xlsx'
    s_path = r'C:\Users\C5293427\Desktop\MA\Simple_Tracker_V3.xlsx'
    EDM_file = EDM(t_path)
    clean = EDM_file.clean_date()
    c = Calender(clean, s_path)
    c.to_xlsx()
    s = Simple_tracker()
    s.create_excel()
   