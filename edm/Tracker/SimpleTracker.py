'''
@Description: SimpleTracker类也是RequestTracker类的子类，它所提供的服务非常简单：(1) 提供一个日历形式的dataframe; (2) 提供一个日历形式的excel表格
@Author: FlyingRedPig
@Date: 2020-04-30 18:03:27
@LastEditors: FlyingRedPig
@LastEditTime: 2020-07-29 12:01:50
@FilePath: \EDM_project\EDM\edm\Tracker\SimpleTracker.py
'''

import sys
sys.path.append("../..")
from edm.Tracker.RequestTracker import Request_Tracker
from edm.Tracker.Analytics import Analytics
import datetime as dt
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Font, colors, Alignment, Border, Side, PatternFill
import time


class SimpleTracker(Request_Tracker):

    def __init__(self, path:str = "../../files/" , *args):
        super().__init__(*args)
        self.__startDate = dt.date.today() + dt.timedelta(-21)
        self.__savePath = path + "Simple_Tracker.xlsx"
        self.__workbook = Workbook()
        self.__wsTitle1 = "Campaign Calender"

    def getStartDate(self):
        return self.__startDate

    def getWorkbook(self):
        return self.__workbook

    def getWsTitle1(self):
        return self.__wsTitle1

    def getSavePath(self):
        return self.__savePath

    def setWorkbook(self, workbook):
        self.__workbook = workbook
        return

    def setStartDate(self, startDate: dt.date):
        self.__startDate = startDate
        return

    def setSavePath(self, savePath):
        self.__savePath = savePath
        return

    def __dateRange(self):
        '''
        @description: 返回simpletracker的开始日期和结束日期
        @param {type} 
        @return: 
        '''
        df = self.getCleanDf()
        df[['Launch Date', 'Event Date']].dropna(inplace=True)
        launchDateMax = df['Launch Date'][df['Launch Date'].apply(
            lambda x: isinstance(x, dt.date))].max()
        eventDateMax = df['Event Date'][df['Event Date'].apply(
            lambda x: isinstance(x, dt.date))].max()

        return pd.date_range(self.getStartDate(),
                             max(launchDateMax, eventDateMax))

    def simpleTrackerDf(self):
        '''
        @description: 向外的接口提供simple tracker dataframe， 实现方式是通过(1)给定date_range; (2)pd.merge
        @param {type} 
        @return: dataframe
        '''
        cleanDf = self.getCleanDf()[[
            'Launch Date', 'Weekday', 'Campaign Name', 'Event Date', 'Owner '
        ]]
        timeSeries = pd.Series(
            self.__dateRange(), name="Launch Date").apply(lambda x: x.date())

        df = pd.merge(
            cleanDf, timeSeries, how="right", on="Launch Date", sort=False)

        df.sort_values(by="Launch Date", ascending=True, inplace=True)
        df['Weekday'] = df['Launch Date'].apply(lambda x: self.turnWeekday(x))
        return df

    def __df2Excel(self):
        '''
        @description: 辅助函数，DataFrame -> openpyxl.workbook
        @param {type} 
        @return: openpyxl.workbook
        '''

        wb = self.getWorkbook()
        ws1 = wb.active
        df1 = self.simpleTrackerDf()
        ws1.title = self.getWsTitle1()
        for r in dataframe_to_rows(df1, index=True, header=True):
            ws1.append(r)

        ws1.delete_cols(1, 1)
        ws1.delete_rows(2, 1)

        return wb

    def __setColumnWidth(self, ws):
        '''
        @description: 设置列宽
        @param {ws:worksheet} 
        @return: void
        '''
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 25
        return

    def __setFont(self, ws):
        '''
        设置字体
        :return:
        '''
        engFont = Font(name='Cambria', size=11, bold=False)
        cnFont = Font(name=u'微软雅黑', size=11, bold=False)
        firstRowFont = Font(name='Times New Roman', size=11, bold=True)

        for col in ws.columns:
            for i in col:
                i.font = engFont

        for i in ws['1']:
            i.font = firstRowFont

        return

    def __setAlign(self, ws):
        '''
        @description: 设置对齐方式
        @param {ws} 
        @return: void
        '''
        align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        for col in ws.columns:
            for i in col:
                i.alignment = align
        return

    def __setBorder(self, ws):
        '''
        @description: 设置边框
        @param {type} 
        @return: void
        '''
        border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000'))

        for col in ws.columns:
            for i in col:
                i.border = border
        return

    def __setColor(self, ws):
        '''
        设置背景颜色
        :return:
        '''
        color_list = ['aecdc2', 'f0b8b8']
        color = 'aecdc2'

        def switch(color, color_list):
            '''
            辅助函数
            '''
            for i in color_list:
                if i == color:
                    continue
                return i

        for i in ws.rows:
            for j in i:
                j.fill = PatternFill('solid', fgColor=color)
            if i[1].value == '星期日':
                color = switch(color, color_list)

        pattern2 = PatternFill('solid', fgColor='f3e8c2')  #header颜色

        for i in ws['1']:
            i.fill = pattern2

        #开始做今天起倒退五天的渐进色

        def oneDayColor(ws, color, date):

            pattern3 = PatternFill('solid', fgColor=color)
            for i in ws.rows:
                if type(i[0].value) == str:
                    continue
                else:
                    if i[0].value.strftime('%Y-%m-%d') == str(date):
                        for j in i:
                            j.fill = pattern3

        today = dt.date.today()
        trafficRange = list(pd.date_range(end=today, periods=5))  #整体转换为列表类型
        trafficRange = list(map(lambda x: x.date(), trafficRange))  #元素转换为date类型
        gradientColor = ['ff6e54', 'ffa600', 'ffa600', 'ffa600', 'ffa600']
        colorDict = dict(zip(trafficRange, gradientColor))  #将日期和颜色对应成字典
        for date in colorDict.keys():
            oneDayColor(ws, colorDict[date], date)

        return

    def __setMerge(self, ws):
        '''
        @description: 合并单元格操作
        @param {type} 
        @return: void
        '''

        col = ws['A']
        date_dic = {}
        index = 0
        start = time.time()
        for i in col:
            index += 1
            item = i.value
            if item not in date_dic.keys():
                date_dic[item] = [index]
            else:
                date_dic[item].append(index)

        end = time.time()

        def merge(index_list):
            ws.merge_cells(
                start_row=min(index_list),
                start_column=1,
                end_row=max(index_list),
                end_column=1)

            ws.merge_cells(
                start_row=min(index_list),
                start_column=2,
                end_row=max(index_list),
                end_column=2)
            return

        #去除所有不需要合并的对象
        index = list(date_dic.values())
        trash = []
        for i in index:
            if len(i) == 1:
                trash.append(i)

        for i in trash:
            index.remove(i)

        list(map(merge, index))  # 对每个日期都合并单元格

        return

    def __setFreeze(self, ws):
        '''
        @description:冻结首行 
        @param {type} 
        @return: void
        '''
        ws.freeze_panes = 'A2'
        return

    def simpleTrackerExcel(self):
        '''
        @description: SimpleTracker类的第二个public接口，通过它更新simple tracker excel
        @param {type} 
        @return: 
        '''

        wb = self.__df2Excel()  #0.019s
        ws = wb[self.getWsTitle1()]  #0s
        self.__setColumnWidth(ws)  #0s
        self.__setFont(ws)  #0.018s
        self.__setAlign(ws)  #0.007s
        self.__setBorder(ws)  #0.037s
        self.__setColor(ws)  #0.07s
        self.__setMerge(ws)  #0.14s
        self.__setFreeze(ws)
        while True:
            try:
                wb.save(self.getSavePath())  #0.03s
                break
            except PermissionError:
                answer = input("simple_tracker.xlsx已打开，请关闭后重试，谢谢(y/n)")
                if answer.lower() == 'y':
                    pass
                elif answer.lower() == 'n':
                    print('由于您执意不关闭simple_tracker.xlsx，我们只好不干活了，结束程序~')
                    break
                else:
                    print('请输入正确的命令(y/n)')

        return


if __name__ == "__main__":
    start = time.time()
    s = SimpleTracker()
    s.simpleTrackerExcel()
    end = time.time()
    print(end - start)
